import openpyxl

xlfile = openpyxl.open("./example.xlsx")


def excelRead():
    #
    print(type(xlfile))
    # 获取工作表
    sheets = xlfile.get_sheet_names()
    sheet1 = xlfile.get_sheet_by_name("Sheet1")
    print(sheet1["A1"].value)

     
    # 获取表的高度和宽度
    max_row = sheet1.max_row
    max_column = sheet1.max_column
    for i in range(1, sheet1.max_row + 1, 2):
        print(i, sheet1.cell(row=i, column=2).value)
     
     

def get_all_rows():
    # 获取工作表
    sheets = xlfile.get_sheet_names()
    sheet1 = xlfile.get_sheet_by_name("Sheet1")
    # 怎么才能获取表中所有元素？
    for row in sheet1.rows:
        for i in row:
            # is this correct?
            print(i.value, end=" ")
        print()


# get_all_rows()
excelRead()
