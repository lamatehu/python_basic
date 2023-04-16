import openpyxl

xml = openpyxl.Workbook()
sheet1 = xml.active
# 设置三个表 sheet1 sheet2 sheet3 sheet1的title是 this one
sheet1.title = "this one"
sheet2 = xml.create_sheet("sheet2")
sheet2.title = "this two"
sheet3 = xml.create_sheet("sheet3")
sheet3.title = "this three"
sheets = xml.get_sheet_names()
sheet1["A1"] = "hello"
sheet1["A2"] = "world"

for row in range(1, sheet1.max_row + 1):
    print(sheet1.cell(row, 1).value, sheet1.cell(row, 2).value)
xml.save("./test.xlsx")
