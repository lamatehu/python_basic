# 导入文件
import os.path

import openpyxl
from openpyxl import workbook

workbook = workbook.Workbook()
sheet1 = workbook.active
for i in range(1, 11):
    sheet1.cell(1, i).value = i-1
for row in range(1, 11):
    sheet1.cell(row, 1).value = row-1

for row in range(2, 11):
    for cel in range(2, 11):
        x = sheet1.cell(1, cel).value
        y = sheet1.cell(row, 1).value
        print(x," ",y)
        strll = f'{x}*{y} = {int(x) * int(y)}'
        sheet1.cell(row, cel).value = strll

if os.path.exists("./99.xlsx"):
    os.remove("./99.xlsx")
sheet1.title = "九九乘法表"
workbook.save("./99.xlsx")
print("done")