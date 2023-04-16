import openpyxl
import pprint

xlsx = openpyxl.load_workbook("./censuspopdata.xlsx")
sheet = xlsx.active
state = ""
country = ""
country_data = {}

for row in range(2, 1000, 1):
    state = sheet.cell(row, 2).value
    country = sheet.cell(row, 3).value
    pop = sheet.cell(row, 4).value
    country_data.setdefault(state, {})
    country_data[state].setdefault(country, {"tracts": 0, "pop": 0})

    country_data[state][country]["tracts"] += 1
    country_data[state][country]["pop"] += int(pop)
    # print(state, country, pop)
    # print(country_data)


print("稍等")
result = open("./censuspopdata.py", "w")
result.write("all_data=" + pprint.pformat(country_data))
result.close()
