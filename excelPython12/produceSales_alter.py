from itertools import product
import logging
import openpyxl

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

PRICE_UPDATES = {"Garlic": 3.07, "Celery": 1.19, "Lemon": 1.27}
xml = openpyxl.load_workbook("./produceSales.xlsx")
sheet1 = xml.active

for row in range(2, sheet1.max_row + 1):
    product_l = sheet1.cell(row, 1).value
    if product_l in PRICE_UPDATES:
        logging.info("出现要修改的信息")
        sheet1.cell(row, 2).value = PRICE_UPDATES[product_l]
        logging.info("修改后的价栏为：%s", sheet1.cell(row, 2).value)
        total = int(sheet1.cell(row, 3).value) * PRICE_UPDATES[product_l]
        sheet1.cell(row, 4).value = total
        logging.info("修改后的总价为：%s", sheet1.cell(row, 4).value)
workbook = openpyxl.Workbook()
workbook.save("./produceSales_alter.xlsx")
