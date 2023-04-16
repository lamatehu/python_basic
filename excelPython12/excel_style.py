import os
import openpyxl
from openpyxl.styles import Font, colors, Alignment, PatternFill, NamedStyle

# 创建洋红色的填充样式
magenta_fill = PatternFill(start_color='FF00FF', end_color='FF00FF', fill_type='solid')
# 创建橘红色的填充样式
orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
# 创建自定义样式
my_style = NamedStyle(name="my_style")
my_style.font = Font(size=12, bold=True, color="FF0000")
my_style.alignment = Alignment(horizontal='center', vertical='center')

# 创建工作簿和工作表
workbook = openpyxl.Workbook()
sheet1 = workbook.active
sheet1.title = "test"

# 设置单元格属性
sheet1.row_dimensions[1].height = 40
sheet1.merge_cells("A1:A4")
sheet1["A1"].value = "test_suan"
sheet1["A1"].fill = magenta_fill
sheet1["A1"].style = my_style

sheet1.merge_cells("D2:D11")
sheet1["D2"].value = "学习不停止"
sheet1["D2"].fill = orange_fill
sheet1["D2"].style = my_style

# 填充数据
i = 1
for row in range(2, 10):
    sheet1.cell(row, 2).value = i
    i += 1

sheet1["A11"].value = "total"
sheet1["B11"].value = "=SUM(B2:B10)"
sheet1.row_dimensions[5].height = 50
# 保存文件
if os.path.exists("./ee.xlsx"):
    os.remove("./ee.xlsx")
workbook.save("./ee.xlsx")

sheet1['5'].height_row = 50